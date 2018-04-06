import openpyxl
import glob

# INIT
OUTPUT_FILENAME = "output.xlsx"
INPUT_FOLDER = "input/"  # Folder containing only input xlsx files
COPY_FROM_ROW = 3  # Copy from row (starting from 1)...
COPY_TO_ROW = 15  # to row...

print("Collating files in %s to file %s, copying rows from %d to %d" %
      (INPUT_FOLDER, OUTPUT_FILENAME, COPY_FROM_ROW, COPY_TO_ROW))

output = openpyxl.Workbook()
sheet_output = output.active
last_row_written = 1

for f in sorted(glob.glob(INPUT_FOLDER+"*.xlsx")):
    wb = openpyxl.load_workbook(filename=f, data_only=True)
    sheet = wb.get_active_sheet()
    if COPY_TO_ROW > sheet.max_row:
        print("  Warning: file %s has only %d rows, hence I have only copied rows from %d to %d" %
              (f, sheet.max_row, COPY_FROM_ROW, sheet.max_row))
    for row in range(COPY_FROM_ROW, min(COPY_TO_ROW, sheet.max_row)+1, 1):
        for col in range(1, sheet.max_column+1, 1):
            sheet_output.cell(row=last_row_written, column=col).value = sheet.cell(row=row, column=col).value
        last_row_written += 1

output.save(filename=OUTPUT_FILENAME)