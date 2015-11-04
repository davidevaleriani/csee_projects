import cherrypy
from cherrypy.lib import auth_basic
import os
import csv
import zipfile
import tempfile
import requests
from requests_ntlm import HttpNtlmAuth
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
import xlrd  # Support to old xls files
from sharepoint import SharePointSite, basic_auth_opener

USERS = {'csee': 'csee'}
marks_dir = "marks/"


class HomePage(object):
    def index(self):
        return """
            <html>
            <head>
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/css/bootstrap.min.css">
                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/js/bootstrap.min.js"></script>
                <title>CSEE Grades Automator</title>
                <style>
                /* Wrapper for page content to push down footer */
                html, body { 
                  padding-top: 0px;
                  font-size: 15px;
                  height: 100%;
                }

                #wrap {
                  min-height: 100%;
                  height: auto !important;
                  height: 100%;
                  /* Negative indent footer by it's height */
                  margin: 0 auto -40px;
                  padding-bottom: 10px;
                }

                /* Set the fixed height of the footer here */
                #push,
                #footer {
                  height: 40px;
                  padding-top: 10px;
                }
                #footer {
                  font-size: 12px;
                  background-color: #ddd;
                  text-align: center;
                }
                </style>
            </head>
            <body>
            <div id="wrap">
                <div class="container">
                    <h2 align="center">Welcome to CSEE Grades Automator</h2>
                    <p align="center">This application allows you to generate the mark forms for final year projects.</p>
                    <p align="center">If you want to download the marks from a zip file <a href="show_marks_report_page">click here</a></p>
                    <p align="center">If you want to download the marks from SharePoint <a href="connect_to_sharepoint">click here</a></p>
                    <br>
                    <form method="get" action="generate" enctype="multipart/form-data" class="form-horizontal">
                        <div class="form-group">
                            <label for="template_sup" class="col-xs-2 col-xs-offset-2 control-label">Template Supervisor</label>
                            <div class="col-xs-6">
                                <input type="file" name="template_sup" id="template_sup" />
                            </div>
                          </div>
                          <div class="form-group">
                            <label for="template_sec" class="col-xs-2 col-xs-offset-2 control-label">Template Second Assessor</label>
                            <div class="col-xs-6">
                                <input type="file" name="template_sec" id="template_sec" />
                            </div>
                          </div>
                        <div class="form-group">
                            <label for="students_list" class="col-xs-2 col-xs-offset-2 control-label">Students' list</label>
                            <div class="col-xs-6">
                                <input type="file" name="students_list" id="students_list" />
                            </div>
                        </div>
                        <div class="form-group">
                            <label for="name_column" class="col-xs-2 col-xs-offset-2 control-label">Student first name column</label>
                            <div class="col-xs-6">
                                <input type="number" class="form-control" name="name_column" min="1" value="3" id="name_column" />
                            </div>
                        </div>
                        <div class="form-group">
                            <label for="name_column" class="col-xs-2 col-xs-offset-2 control-label">Student surname column</label>
                            <div class="col-xs-6">
                                <input type="number" class="form-control" name="surname_column" min="1" value="2" id="surname_column" />
                            </div>
                        </div>
                        <div class="form-group">
                            <label for="regno_column" class="col-xs-2 col-xs-offset-2 control-label">Student registration number column</label>
                            <div class="col-xs-6">
                                <input type="number" class="form-control" name="regno_column" min="1" value="1" id="regno_column" />
                            </div>
                        </div>
                        <div class="form-group">
                            <label for="sup_column" class="col-xs-2 col-xs-offset-2 control-label">Student supervisor column</label>
                            <div class="col-xs-6">
                                <input type="number" class="form-control" name="sup_column" min="1" value="7" id="sup_column" />
                            </div>
                        </div>
                        <div class="form-group">
                            <label for="sec_column" class="col-xs-2 col-xs-offset-2 control-label">Student second assessor column</label>
                            <div class="col-xs-6">
                                <input type="number" class="form-control" name="sec_column" min="1" value="8" id="sec_column" />
                            </div>
                        </div>
                        <div class="form-group">
                            <div class="col-xs-offset-6 col-sm-6">
                                <button type="submit" class="btn btn-primary">Generate</button>
                            </div>
                        </div>
                    </form>
                </div>
                <div id="push"></div>
            </div>
            <div id="footer">
                <div class="container">
                    <p class="muted credit">Created by <a href="http://www.davidevaleriani.it">Davide Valeriani</a>.</p>
                </div>
            </div>
            </body>
            </html>
            """
    index.exposed = True

    @cherrypy.expose
    def generate(self, template_sup="template_supervisor.xlsx", template_sec="template_second_assessor.xlsx", students_list="list_of_students.xls", name_column=2, surname_column=3, regno_column=1, sup_column=7, sec_column=8):
        # Load the template
        if template_sup.split(".")[-1].upper() == "XLSX" and template_sec.split(".")[-1].upper() == "XLSX":
            mark_form_sup_doc = load_workbook(filename=template_sup)
            first_sheet = mark_form_sup_doc.get_sheet_names()[0]
            mark_form_sup = mark_form_sup_doc.get_sheet_by_name(first_sheet)
            mark_form_sec_doc = load_workbook(filename=template_sec)
            first_sheet = mark_form_sec_doc.get_sheet_names()[0]
            mark_form_sec = mark_form_sec_doc.get_sheet_by_name(first_sheet)
            border = Border(left=Side(border_style='thin', color='00000000'),
                            right=Side(border_style='thin', color='00000000'),
                            top=Side(border_style='thin', color='00000000'),
                            bottom=Side(border_style='thin', color='00000000'),
                            vertical=Side(border_style='thin', color='00000000'),
                            horizontal=Side(border_style='thin', color='00000000'))

        else:
            return """
            <html>
            <head>
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/css/bootstrap.min.css">
                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/js/bootstrap.min.js"></script>
                <title>CSEE Grades Automator</title>
                <style>
                /* Wrapper for page content to push down footer */
                html, body { 
                  padding-top: 0px;
                  font-size: 15px;
                  height: 100%;
                }

                #wrap {
                  min-height: 100%;
                  height: auto !important;
                  height: 100%;
                  /* Negative indent footer by it's height */
                  margin: 0 auto -40px;
                  padding-bottom: 10px;
                }

                /* Set the fixed height of the footer here */
                #push,
                #footer {
                  height: 40px;
                  padding-top: 10px;
                }
                #footer {
                  font-size: 12px;
                  background-color: #ddd;
                  text-align: center;
                }
                </style>
            </head>
            <body>
            <div id="wrap">
                <div class="container">
                    <h2 align="center" style="color:red">You have to select a xlsx file for the template.</h2>
                    <div align="center">
                        <a href="index" class="btn btn-default">Back</a>
                    </div>
                </div>
                <div id="push"></div>
            </div>
            <div id="footer">
                <div class="container">
                    <p class="muted credit">Created by <a href="http://www.davidevaleriani.it">Davide Valeriani</a>.</p>
                </div>
            </div>
            </body>
            </html>
            """
            raise Exception("Only xlsx files are supported so far for templates")

        # Load the students list
        if students_list.split(".")[-1].upper() == "XLS":
            workbook = xlrd.open_workbook(students_list)
            worksheets = workbook.sheet_names()
            worksheet = workbook.sheet_by_name(worksheets[0])
            # Skip the header row
            for row in range(1, worksheet.nrows):
                student_surname = worksheet.cell_value(row, int(surname_column))
                student_name = worksheet.cell_value(row, int(name_column))
                student_regno = worksheet.cell_value(row, int(regno_column))
                student_sup = worksheet.cell_value(row, int(sup_column))
                student_sec = worksheet.cell_value(row, int(sec_column))
                # Populating
                # Name
                mark_form_sup['C3'] = student_name+" "+student_surname
                mark_form_sec['C3'] = student_name+" "+student_surname
                # Registration number
                mark_form_sup['C4'] = student_regno
                mark_form_sec['C4'] = student_regno
                # Supervisor
                mark_form_sup['C5'] = mark_form_sec['C5'] = student_sup
                # Second assessor
                mark_form_sup['C6'] = mark_form_sec['C6'] = student_sec
                # Fixing borders
                for r in mark_form_sup['G3':'G6']:
                    for cell in r:
                        cell.border = border.copy()
                for r in mark_form_sup['J2':'J44']:
                    for cell in r:
                        cell.border = border.copy()
                for r in mark_form_sup['L1':'L45']:
                    for cell in r:
                        cell.border = Border(left=Side(border_style='medium', color='00000000'))
                for r in mark_form_sup['A10':'A44']:
                    for cell in r:
                        cell.border = border.copy()
                for r in mark_form_sup['A46':'K46']:
                    for cell in r:
                        cell.border = Border(top=Side(border_style='medium', color='00000000'))
                for r in mark_form_sec['F1':'F33']:
                    for cell in r:
                        cell.border = border.copy(right=Side(border_style='medium', color='00000000'))
                for r in mark_form_sec['C3':'E6']:
                    for cell in r:
                        cell.border = border.copy()
                for r in mark_form_sec['B8':'E8']:
                    for cell in r:
                        cell.border = border.copy()
                for r in mark_form_sec['A9':'A32']:
                    for cell in r:
                        cell.border = border.copy()
                mark_form_sup['H27'].border = border.copy()
                mark_form_sup['I27'].border = border.copy()
                mark_form_sup['H37'].border = border.copy()
                mark_form_sup['I37'].border = border.copy()
                mark_form_sec['F1'].border = Border(right=Side(border_style='medium', color='00000000'))
                mark_form_sec['F2'].border = border.copy(right=Side(border_style='medium', color='00000000'),
                                                         top=Side(border_style=None),
                                                         left=Side(border_style=None))
                mark_form_sec['F7'].border = border.copy(right=Side(border_style='medium', color='00000000'),
                                                         left=Side(border_style=None))
                mark_form_sec['F8'].border = border.copy(right=Side(border_style='medium', color='00000000'))
                mark_form_sec['F14'].border = border.copy(right=Side(border_style='medium', color='00000000'))
                mark_form_sec['F21'].border = border.copy(right=Side(border_style='medium', color='00000000'))
                mark_form_sec['F27'].border = border.copy(right=Side(border_style='medium', color='00000000'))
                mark_form_sec['D26'].border = border.copy()
                mark_form_sec['D20'].border = border.copy()
                mark_form_sec['D13'].border = border.copy()
                mark_form_sec['F33'].border = border.copy(right=Side(border_style='medium', color='00000000'),
                                                          bottom=Side(border_style='medium', color='00000000'),)

                # Saving files in the specific folders
                if not os.path.isdir(marks_dir):
                    os.makedirs(marks_dir)
                if not os.path.isdir(marks_dir+student_sup+"/"):
                    os.makedirs(marks_dir+student_sup+"/")
                mark_form_sup_doc.save(marks_dir+student_sup+"/"+student_surname+'_sup.xlsx')
                if not os.path.isdir(marks_dir+student_sec+"/"):
                    os.makedirs(marks_dir+student_sec+"/")
                mark_form_sec_doc.save(marks_dir+student_sec+"/"+student_surname+'_sec.xlsx')

        elif students_list.split(".")[-1].upper() == "XLSX":
            workbook = load_workbook(students_list, use_iterators=True)
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet)
            row_iterator = worksheet.iter_rows()
            # Skip the first row
            next(row_iterator)
            for row in row_iterator:
                student_names = row[int(name_column)-1].value
                student_regno = row[int(regno_column)-1].value
                student_sup = row[int(sup_column)-1].value
                student_sec = row[int(sec_column)-1].value
                try:
                    tmp = int(student_regno)
                except:
                    break
                try:
                    student_surname = student_names.split(",")[0].strip(" ")
                    student_name = student_names.split(",")[1].strip(" ")
                except:
                    # Handle the cases where there is no comma in the name cell
                    student_surname = student_names.split(" ")[0].strip(" ")
                    student_name = student_names.split(" ")[1].strip(" ")
                # Populating
                # Name
                mark_form_sup['C3'] = student_name+" "+student_surname
                mark_form_sec['C3'] = student_name+" "+student_surname
                # Registration number
                mark_form_sup['C4'] = student_regno
                mark_form_sec['C4'] = student_regno
                # Supervisor
                mark_form_sup['C5'] = mark_form_sec['C5'] = student_sup
                # Second assessor
                mark_form_sup['C6'] = mark_form_sec['C6'] = student_sec
                # Saving files in the specific folders
                if not os.path.isdir(marks_dir):
                    os.makedirs(marks_dir)
                if not os.path.isdir(marks_dir+student_sup+"/"):
                    os.makedirs(marks_dir+student_sup+"/")
                mark_form_sup_doc.save(marks_dir+student_sup+"/"+student_surname+'_sup.xlsx')
                if not os.path.isdir(marks_dir+student_sec+"/"):
                    os.makedirs(marks_dir+student_sec+"/")
                mark_form_sec_doc.save(marks_dir+student_sec+"/"+student_surname+'_sec.xlsx')
        else:
            raise Exception("Unable to serve this type of file")
        
        print "Generate zip"
        # Zip the marks
        zipf = zipfile.ZipFile('marks.zip', 'w')
        for root, dirs, files in os.walk(marks_dir):
            for f in files:
                zipf.write(os.path.join(root, f))
        zipf.close()
        # Remove the marks directory
        shutil.rmtree(marks_dir)
        # Return the file to download
        marks_file = open("marks.zip", 'r')
        return cherrypy.lib.static.serve_fileobj(marks_file, disposition='attachment', name="marks.zip")

    @cherrypy.expose
    def get_marks(self):
        marks_file = open("marks.zip", 'r')
        return cherrypy.lib.static.serve_fileobj(marks_file, disposition='attachment', name="marks.zip")

    @cherrypy.expose
    def connect_to_sharepoint(self):
        # sudo pip install sharepoint
        server_url = "https://sp.essex.ac.uk/"
        site_url = server_url+"depts/csee/Pages/Default.aspx"
        username = "dvaler"
        password = ""
        headers = {'accept': 'application/json;odata=verbose'}
        r = requests.get("https://sp.essex.ac.uk/depts/csee/_api/web/getfolderbyserverrelativeurl('/depts/csee/Administration%20201516/Assessment/CE301')", auth=HttpNtlmAuth('CAMPUS\\'+username, password), headers=headers)

    @cherrypy.expose
    def show_marks_report_page(self):
        return """
            <html>
            <head>
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/css/bootstrap.min.css">
                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.4/js/bootstrap.min.js"></script>
                <title>CSEE Grades Automator</title>
                <style>
                /* Wrapper for page content to push down footer */
                html, body {
                  padding-top: 0px;
                  font-size: 15px;
                  height: 100%;
                }

                #wrap {
                  min-height: 100%;
                  height: auto !important;
                  height: 100%;
                  /* Negative indent footer by it's height */
                  margin: 0 auto -40px;
                  padding-bottom: 10px;
                }

                /* Set the fixed height of the footer here */
                #push,
                #footer {
                  height: 40px;
                  padding-top: 10px;
                }
                #footer {
                  font-size: 12px;
                  background-color: #ddd;
                  text-align: center;
                }
                </style>
            </head>
            <body>
            <div id="wrap">
                <div class="container">
                    <h2 align="center">Welcome to CSEE Grades Automator</h2>
                    <p align="center">This application allows you to download the marks from a zip file in a spreadsheet.</p>
                    <p align="center">If you want to generate the marks forms <a href="index">click here</a></p>
                    <p align="center">If you want to download the marks from SharePoint <a href="connect_to_sharepoint">click here</a></p>
                    <br>
                    <form method="post" action="get_marks_report" enctype="multipart/form-data" class="form-horizontal">
                        <div class="form-group">
                            <label for="marks" class="col-xs-2 col-xs-offset-2 control-label">Zip file</label>
                            <div class="col-xs-6">
                                <input type="file" name="marks" id="marks" />
                            </div>
                        </div>
                        <div class="form-group">
                            <div class="col-xs-offset-6 col-sm-6">
                                <button type="submit" class="btn btn-primary">Generate</button>
                            </div>
                        </div>
                    </form>
                </div>
                <div id="push"></div>
            </div>
            <div id="footer">
                <div class="container">
                    <p class="muted credit">Created by <a href="http://www.davidevaleriani.it">Davide Valeriani</a>.</p>
                </div>
            </div>
            </body>
            </html>
            """

    @cherrypy.expose
    def get_marks_report(self, marks):
        if not os.path.isdir("tmp"):
            os.mkdir("tmp")
        # Extract files
        if type(marks) != file and hasattr(marks, "file"):
            marks = marks.file
        zipf = zipfile.ZipFile(marks, 'r')
        zipf.extractall("tmp/")
        # Init the marks spreadsheet
        marks_wb = Workbook()
        marks_ws = marks_wb.active
        # Header
        marks_ws['A1'] = "Student first name"
        marks_ws['B1'] = "Student surname"
        marks_ws['C1'] = "Student registration number"
        marks_ws['D1'] = "Supervisor first name"
        marks_ws['E1'] = "Supervisor surname"
        marks_ws['F1'] = "2nd assessor first name"
        marks_ws['G1'] = "2nd assessor surname"
        marks_ws['H1'] = "Initial report mark"
        marks_ws['I1'] = "Interim report mark"
        marks_ws['J1'] = "Poster mark"
        marks_ws['K1'] = "Final report mark"
        marks_ws['L1'] = "Logbook mark"
        marks_ws['M1'] = "PDO mark"
        marks_ws['N1'] = "Module total"

        # For each mark form
        for dirname, dirnames, filenames in os.walk('tmp/'):
            for f in filenames:
                if "_sup" in f:
                    # Get the marks
                    print dirname+"/"+f
                    doc = load_workbook(filename=dirname+"/"+f)
                    first_sheet = doc.get_sheet_names()[0]
                    form = doc.get_sheet_by_name(first_sheet)
                    student_name = " ".join(form["C3"].value.split(" ")[:-1])
                    student_surname = form["C3"].value.split(" ")[-1]
                    student_regno = form["C4"].value
                    sup_name = form["C5"].value.split()[0]
                    sup_surname = form["C5"].value.split()[1]
                    sec_name = form["C6"].value.split()[0]
                    sec_surname = form["C6"].value.split()[1]
                    initial_report_mark = self.multiply_and_sum(form, 11, 13)
                    interim_report_mark = self.multiply_and_sum(form, 17, 20)
                    poster_mark = self.multiply_and_sum(form, 24, 26, column1="C")
                    final_report_mark = self.multiply_and_sum(form, 30, 32)
                    logbook_mark = self.multiply_and_sum(form, 36, 36, column1="C")
                    pdo_mark = self.multiply_and_sum(form, 40, 43)
                    total = initial_report_mark*0.05+interim_report_mark*0.20+poster_mark*0.05+final_report_mark*0.5+logbook_mark*0.05+pdo_mark*0.15
                    marks_ws.append([student_name, student_surname, student_regno, sup_name, sup_surname, sec_name, sec_surname, initial_report_mark, interim_report_mark, poster_mark, final_report_mark, logbook_mark, pdo_mark, total])
                else:
                    continue
        # Save the report
        marks_wb.save("marks_report.xlsx")

        # Remove the marks
        shutil.rmtree('tmp')

        # Return the file to download
        marks_report = open("marks_report.xlsx", 'r')
        return cherrypy.lib.static.serve_fileobj(marks_report, disposition='attachment', name="marks_report.xlsx")

    def multiply_and_sum(self, sheet, first_row, last_row, column1="E", column2="F"):
        mark = 0
        for r in range(first_row, last_row+1):
            try:
                mark += sheet[column1+str(r)].value*sheet[column2+str(r)].value
            except:
                if sheet[column1+str(r)].value[0] == "=":
                    mark += 0.5*(sheet[chr(ord(column1)-2)+str(r)].value+sheet[chr(ord(column1)-1)+str(r)].value)*sheet[column2+str(r)].value
                else:
                    print column1+str(r), column2+str(r), sheet[column1+str(r)].value, sheet[column2+str(r)].value
                    raise ValueError()
        return mark


if __name__ == '__main__':
    cherrypy.server.socket_host = '0.0.0.0'
    cherrypy.quickstart(HomePage())
